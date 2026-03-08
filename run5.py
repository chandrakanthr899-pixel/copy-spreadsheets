import os
import shutil
import time
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


def log_step(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def backup_file(path: str) -> str:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    base = os.path.basename(path)
    name, ext = os.path.splitext(base)

    ts = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{name}-{ts}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_sheet_headers_openpyxl(file_path: str, sheet_name: str, header_row: int = 2) -> dict:
    """
    Reads only the header row from a sheet and returns:
    {
        "ColumnName": column_index
    }
    """
    log_step(f"Reading headers from '{sheet_name}' in {os.path.basename(file_path)}")
    wb = load_workbook(file_path, read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        ws = wb[sheet_name]
        headers = {}

        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
            for idx, cell_value in enumerate(row, start=1):
                header = normalize_header(cell_value)
                if header:
                    headers[header] = idx
            break

        if not headers:
            raise ValueError(
                f"No headers found in sheet '{sheet_name}' at row {header_row} in {file_path}"
            )

        return headers
    finally:
        wb.close()


def clear_extra_destination_rows(ws, start_row: int, max_row: int, max_col: int, chunk_size: int = 1000) -> None:
    """
    Clears cell values only, preserving formatting.
    """
    if max_row < start_row:
        return

    total_rows = max_row - start_row + 1
    log_step(f"Clearing existing destination data: {total_rows} row(s)")

    with tqdm(total=total_rows, desc="Clearing rows", unit="row") as pbar:
        for chunk_start in range(start_row, max_row + 1, chunk_size):
            chunk_end = min(chunk_start + chunk_size - 1, max_row)
            for row in ws.iter_rows(
                min_row=chunk_start,
                max_row=chunk_end,
                min_col=1,
                max_col=max_col
            ):
                for cell in row:
                    cell.value = None

            pbar.update(chunk_end - chunk_start + 1)


def copy_dataframe_to_sheet(
    df: pd.DataFrame,
    ws,
    dst_header_map: dict,
    data_start_row: int = 3,
    chunk_size: int = 1000,
) -> None:
    """
    Writes DataFrame rows into worksheet by matching column names to destination headers.
    Preserves formatting because only cell values are updated.
    """
    row_count = len(df)
    if row_count == 0:
        log_step("Source DataFrame is empty. Nothing to write.")
        return

    log_step(f"Writing {row_count} row(s) to destination sheet")

    # Precompute destination column indexes in DataFrame order for speed
    df_columns = list(df.columns)
    dst_col_indexes = [dst_header_map[col] for col in df_columns]

    # Convert once for faster iteration than repeated DataFrame access
    records = list(df.itertuples(index=False, name=None))

    with tqdm(total=row_count, desc="Writing rows", unit="row") as pbar:
        for chunk_start in range(0, row_count, chunk_size):
            chunk_end = min(chunk_start + chunk_size, row_count)
            chunk = records[chunk_start:chunk_end]

            for offset, record in enumerate(chunk):
                excel_row = data_start_row + chunk_start + offset

                for value, dst_col in zip(record, dst_col_indexes):
                    ws.cell(row=excel_row, column=dst_col, value=value)

            pbar.update(len(chunk))


def copy_excel_data_by_headers(
    source_file: str,
    destination_file: str,
    source_sheet: str = "SheetA",
    destination_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
    chunk_size: int = 1000,
) -> None:
    start_time = time.time()

    log_step("Creating backups before making changes")
    src_backup = backup_file(source_file)
    dst_backup = backup_file(destination_file)
    log_step(f"Backup created: {src_backup}")
    log_step(f"Backup created: {dst_backup}")

    # Read destination headers first so pandas only loads matching columns
    dst_headers = read_sheet_headers_openpyxl(destination_file, destination_sheet, header_row)
    dst_header_names = list(dst_headers.keys())
    dst_header_set = set(dst_header_names)

    # Read source using pandas
    log_step(f"Loading source data from '{source_sheet}' using pandas")
    # header_row is Excel row number; pandas header is zero-based
    df_source = pd.read_excel(
        source_file,
        sheet_name=source_sheet,
        header=header_row - 1,
        dtype=object,
        engine="openpyxl",
    )

    # Normalize source column names
    df_source.columns = [normalize_header(col) for col in df_source.columns]

    # Keep only columns that exist in destination
    common_columns = [col for col in df_source.columns if col in dst_header_set]

    if not common_columns:
        raise ValueError("No matching column names found between source and destination sheets.")

    missing_in_destination = [col for col in df_source.columns if col and col not in dst_header_set]
    if missing_in_destination:
        log_step(
            "Skipping source columns not found in destination: "
            + ", ".join(missing_in_destination)
        )

    df_source = df_source[common_columns]

    # Optional: remove fully empty rows
    original_row_count = len(df_source)
    df_source = df_source.dropna(how="all").reset_index(drop=True)
    removed_empty_rows = original_row_count - len(df_source)
    if removed_empty_rows > 0:
        log_step(f"Removed {removed_empty_rows} fully empty row(s) from source data")

    log_step(f"Source rows to copy: {len(df_source)}")
    log_step(f"Matching columns: {len(common_columns)}")

    # Open destination workbook in normal mode for writing
    log_step(f"Opening destination workbook '{os.path.basename(destination_file)}'")
    wb = load_workbook(destination_file)
    try:
        if destination_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{destination_sheet}' not found in {destination_file}")

        ws = wb[destination_sheet]

        # Clear only existing data rows, preserve header/formatting
        max_existing_row = ws.max_row
        max_existing_col = ws.max_column

        clear_extra_destination_rows(
            ws=ws,
            start_row=data_start_row,
            max_row=max_existing_row,
            max_col=max_existing_col,
            chunk_size=chunk_size,
        )

        # Write new data
        copy_dataframe_to_sheet(
            df=df_source,
            ws=ws,
            dst_header_map=dst_headers,
            data_start_row=data_start_row,
            chunk_size=chunk_size,
        )

        log_step("Saving destination workbook... this can take time for large files")
        wb.save(destination_file)
        log_step("Save completed")

    finally:
        wb.close()

    elapsed = time.time() - start_time
    log_step(f"Done in {elapsed:.2f} seconds")


if __name__ == "__main__":
    SOURCE_FILE = r"C:\path\to\first.xlsx"
    DESTINATION_FILE = r"C:\path\to\second.xlsx"

    copy_excel_data_by_headers(
        source_file=SOURCE_FILE,
        destination_file=DESTINATION_FILE,
        source_sheet="SheetA",
        destination_sheet="SheetB",
        header_row=2,
        data_start_row=3,
        chunk_size=1000,   # try 2000 or 5000 for larger files
    )