import os
import shutil
from datetime import datetime

from openpyxl import load_workbook
from tqdm import tqdm


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def backup_file(path: str) -> str:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    base = os.path.basename(path)
    name, ext = os.path.splitext(base)

    ts = datetime.now().strftime("%H%M%S")  # hhmmss
    backup_name = f"backup-{name}-{ts}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def read_headers(ws, header_row: int) -> dict:
    """
    Returns dict: {header_name: column_index}
    Ignores empty header cells.
    """
    log(f"Reading headers from sheet '{ws.title}', row {header_row}")
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            headers[key] = col

    log(f"Found {len(headers)} headers in sheet '{ws.title}'")
    return headers


def clear_data_rows_in_batches(ws, start_row: int, batch_size: int = 500):
    """
    Clears values from start_row to ws.max_row (all columns) in batches.
    Keeps formatting; only clears cell values.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < start_row:
        log(f"No destination rows to clear in sheet '{ws.title}'")
        return

    total_rows = max_row - start_row + 1
    log(
        f"Starting clear operation for sheet '{ws.title}': "
        f"{total_rows} rows x {max_col} columns"
    )

    batch_num = 0
    for batch_start in range(start_row, max_row + 1, batch_size):
        batch_num += 1
        batch_end = min(batch_start + batch_size - 1, max_row)

        for r in range(batch_start, batch_end + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).value = None

        log(
            f"Cleared batch {batch_num}: rows {batch_start} to {batch_end} "
            f"({batch_end - batch_start + 1} rows)"
        )


def copy_sheet_data(
    src_path: str,
    dst_path: str,
    src_sheet: str = "SheetA",
    dst_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
    batch_size: int = 500,
):
    log("Starting Excel copy operation")

    # 1) Backups FIRST
    log("Creating backups")
    b1 = backup_file(src_path)
    b2 = backup_file(dst_path)
    log(f"Backup created: {b1}")
    log(f"Backup created: {b2}")

    # 2) Load workbooks
    log(f"Loading source workbook: {src_path}")
    src_wb = load_workbook(src_path, data_only=False)

    log(f"Loading destination workbook: {dst_path}")
    dst_wb = load_workbook(dst_path, data_only=False)

    if src_sheet not in src_wb.sheetnames:
        raise ValueError(f"Source sheet '{src_sheet}' not found in {src_path}")
    if dst_sheet not in dst_wb.sheetnames:
        raise ValueError(f"Destination sheet '{dst_sheet}' not found in {dst_path}")

    src_ws = src_wb[src_sheet]
    dst_ws = dst_wb[dst_sheet]

    log(f"Source sheet loaded: '{src_ws.title}'")
    log(f"Destination sheet loaded: '{dst_ws.title}'")

    # 3) Header maps (row 2)
    src_headers = read_headers(src_ws, header_row)
    dst_headers = read_headers(dst_ws, header_row)

    if not src_headers:
        raise ValueError(f"No headers found in source sheet '{src_sheet}' at row {header_row}")
    if not dst_headers:
        raise ValueError(f"No headers found in destination sheet '{dst_sheet}' at row {header_row}")

    # 4) Common columns only (by name)
    common_cols = [h for h in src_headers.keys() if h in dst_headers]
    missing_in_dst = [h for h in src_headers.keys() if h not in dst_headers]

    log(f"Common columns found: {len(common_cols)}")

    if not common_cols:
        raise ValueError("No matching column names found between source and destination.")

    if missing_in_dst:
        log("These source columns are missing in destination and will be skipped:")
        for h in missing_in_dst:
            log(f"  - {h}")

    # 5) Determine source data range
    last_row = src_ws.max_row
    if last_row < data_start_row:
        log("No data rows found to copy (source has no rows below header).")
        return

    total_rows = last_row - data_start_row + 1
    log(f"Total source data rows to process: {total_rows}")

    # 6) Clear destination data rows (overwrite behavior)
    log("Starting destination clear step")
    clear_data_rows_in_batches(dst_ws, data_start_row, batch_size=batch_size)
    log("Destination clear step completed")

    # 7) Copy row-by-row in batches
    log(f"Starting copy step with batch size {batch_size}")

    batch_num = 0
    with tqdm(total=total_rows, desc="Copying rows", unit="row") as pbar:
        for batch_start in range(data_start_row, last_row + 1, batch_size):
            batch_num += 1
            batch_end = min(batch_start + batch_size - 1, last_row)

            for src_r in range(batch_start, batch_end + 1):
                dst_r = src_r  # assumes copy starts from same data_start_row

                for h in common_cols:
                    src_c = src_headers[h]
                    dst_c = dst_headers[h]
                    dst_ws.cell(row=dst_r, column=dst_c).value = src_ws.cell(
                        row=src_r, column=src_c
                    ).value

            processed = batch_end - batch_start + 1
            pbar.update(processed)

            log(
                f"Copied batch {batch_num}: source rows {batch_start} to {batch_end} "
                f"({processed} rows)"
            )

    # 8) Save destination workbook
    log("Saving destination workbook")
    dst_wb.save(dst_path)
    log(f"Done. Data copied from '{src_sheet}' ({src_path}) to '{dst_sheet}' ({dst_path}).")

    # 9) Close workbooks
    log("Closing workbooks")
    src_wb.close()
    dst_wb.close()
    log("Operation completed successfully")


if __name__ == "__main__":
    # EDIT THESE:
    SRC_FILE = r"source.xlsx"
    DST_FILE = r"destination.xlsx"

    copy_sheet_data(
        src_path=SRC_FILE,
        dst_path=DST_FILE,
        src_sheet="July Template",
        dst_sheet="template",
        header_row=2,
        data_start_row=3,
        batch_size=500,  # try 200 / 500 / 1000 depending on file size
    )