import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


def backup_file(path: str) -> str:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    filename = os.path.basename(path)
    name, ext = os.path.splitext(filename)

    ts = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{name}-{ts}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def get_headers_from_row(ws, header_row: int) -> dict:
    """
    Returns:
        {
            "ColumnName": column_index
        }
    """
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if value is None:
            continue
        key = str(value).strip()
        if key:
            headers[key] = col_idx
    return headers


def clear_old_data(ws, start_row: int, start_col: int, end_col: int, num_rows: int):
    """
    Clear only the existing data block that is relevant.
    This is much better than clearing the whole sheet.
    """
    if num_rows <= 0:
        return

    end_row = start_row + num_rows - 1
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col
    ):
        for cell in row:
            cell.value = None


def copy_data_pandas_openpyxl(
    src_path: str,
    dst_path: str,
    src_sheet: str = "SheetA",
    dst_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
):
    # --------------------------------------------------
    # 1) Backup both files first
    # --------------------------------------------------
    src_backup = backup_file(src_path)
    dst_backup = backup_file(dst_path)

    print(f"Backup created: {src_backup}")
    print(f"Backup created: {dst_backup}")

    # --------------------------------------------------
    # 2) Read source data with pandas (fast bulk read)
    #    header_row=2 means pandas header index = 1
    # --------------------------------------------------
    print("Reading source sheet with pandas...")
    df_src = pd.read_excel(
        src_path,
        sheet_name=src_sheet,
        header=header_row - 1,
        engine="openpyxl"
    )

    # Remove fully empty rows
    df_src = df_src.dropna(how="all")

    # Normalize column names
    df_src.columns = [str(col).strip() for col in df_src.columns]

    if df_src.empty:
        print("No source data found to copy.")
        return

    # --------------------------------------------------
    # 3) Load destination workbook with openpyxl
    #    keep_links=False can reduce overhead for files with links
    # --------------------------------------------------
    print("Loading destination workbook with openpyxl...")
    dst_wb = load_workbook(dst_path, keep_links=False)

    if dst_sheet not in dst_wb.sheetnames:
        raise ValueError(f"Destination sheet '{dst_sheet}' not found in {dst_path}")

    dst_ws = dst_wb[dst_sheet]

    # --------------------------------------------------
    # 4) Read destination headers from row 2
    # --------------------------------------------------
    dst_headers = get_headers_from_row(dst_ws, header_row)
    if not dst_headers:
        raise ValueError(f"No headers found in destination sheet '{dst_sheet}' at row {header_row}")

    # --------------------------------------------------
    # 5) Match common columns
    # --------------------------------------------------
    common_cols = [col for col in df_src.columns if col in dst_headers]
    missing_in_dst = [col for col in df_src.columns if col not in dst_headers]

    if not common_cols:
        raise ValueError("No matching columns found between source and destination.")

    if missing_in_dst:
        print("\nThese source columns are missing in destination and will be skipped:")
        for col in missing_in_dst:
            print(f"  - {col}")

    # Keep only common columns
    df_to_copy = df_src[common_cols].copy()

    # Convert NaN to None so Excel cells are blank instead of NaN
    df_to_copy = df_to_copy.where(pd.notnull(df_to_copy), None)

    # --------------------------------------------------
    # 6) Determine old data size in destination
    #    Only clear the columns we are going to write
    # --------------------------------------------------
    existing_data_rows = max(dst_ws.max_row - data_start_row + 1, 0)
    new_data_rows = len(df_to_copy)

    rows_to_clear = max(existing_data_rows, new_data_rows)

    dst_col_indexes = [dst_headers[col] for col in common_cols]
    min_dst_col = min(dst_col_indexes)
    max_dst_col = max(dst_col_indexes)

    print("Clearing previous destination data block...")
    clear_old_data(
        dst_ws,
        start_row=data_start_row,
        start_col=min_dst_col,
        end_col=max_dst_col,
        num_rows=rows_to_clear
    )

    # --------------------------------------------------
    # 7) Write data row-by-row with progress bar
    #    Faster than reading cell-by-cell from source,
    #    because pandas already did the heavy lifting.
    # --------------------------------------------------
    print("Writing data to destination sheet...")
    row_records = df_to_copy.to_dict(orient="records")

    for row_offset, record in enumerate(
        tqdm(row_records, total=len(row_records), desc="Copying rows")
    ):
        excel_row = data_start_row + row_offset
        for col_name in common_cols:
            dst_col = dst_headers[col_name]
            dst_ws.cell(row=excel_row, column=dst_col, value=record[col_name])

    # --------------------------------------------------
    # 8) Save once
    # --------------------------------------------------
    print("Saving destination workbook...")
    dst_wb.save(dst_path)

    print("\nDone.")
    print(f"Copied {len(df_to_copy)} rows from '{src_sheet}' in source file to '{dst_sheet}' in destination file.")


if __name__ == "__main__":
    SRC_FILE = r"C:\path\to\first.xlsx"
    DST_FILE = r"C:\path\to\second.xlsx"

    copy_data_pandas_openpyxl(
        src_path=SRC_FILE,
        dst_path=DST_FILE,
        src_sheet="SheetA",
        dst_sheet="SheetB",
        header_row=2,
        data_start_row=3,
    )