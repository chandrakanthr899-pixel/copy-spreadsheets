import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


def backup_file(path: str) -> str:
    """
    Creates a backup in the same folder:
    backup-<originalfilename>-HHMMSS.xlsx

    Example:
    report.xlsx -> backup-report-143012.xlsx
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    filename = os.path.basename(path)
    stem, ext = os.path.splitext(filename)

    timestamp = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{stem}-{timestamp}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def get_headers_from_row(ws, header_row: int) -> dict:
    """
    Returns:
        {
            "ColumnName": column_index,
            ...
        }

    Reads only one row from the worksheet.
    """
    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            header = str(cell.value).strip()
            if header:
                headers[header] = cell.column
    return headers


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace pandas NaN/NaT with None so openpyxl writes clean blanks.
    """
    return df.where(pd.notna(df), None)


def clear_extra_rows(ws, start_row: int, end_row: int, col_indexes: list[int], desc: str):
    """
    Clears values only for the given column indexes in the row range.
    Keeps formatting intact.
    """
    if end_row < start_row:
        return

    total_rows = end_row - start_row + 1
    for row_num in tqdm(range(start_row, end_row + 1), total=total_rows, desc=desc):
        for col_idx in col_indexes:
            ws.cell(row=row_num, column=col_idx).value = None


def copy_sheet_data_pandas_openpyxl(
    source_file: str,
    target_file: str,
    source_sheet: str = "SheetA",
    target_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
):
    """
    Hybrid approach:
    - pandas reads source data fast
    - openpyxl updates destination workbook to preserve formatting
    """

    print("Creating backups...")
    src_backup = backup_file(source_file)
    dst_backup = backup_file(target_file)
    print(f"Source backup : {src_backup}")
    print(f"Target backup : {dst_backup}")

    print("\nLoading source data with pandas...")
    # header_row=2 means pandas header index = 1 (0-based)
    df_src = pd.read_excel(
        source_file,
        sheet_name=source_sheet,
        header=header_row - 1,
        dtype=object,
        engine="openpyxl",
    )
    df_src = normalize_dataframe(df_src)

    print(f"Source rows loaded: {len(df_src)}")
    print(f"Source columns loaded: {len(df_src.columns)}")

    print("\nLoading target workbook with openpyxl...")
    wb_target = load_workbook(target_file)
    if target_sheet not in wb_target.sheetnames:
        raise ValueError(f"Target sheet '{target_sheet}' not found in {target_file}")

    ws_target = wb_target[target_sheet]

    print("Reading target headers...")
    target_headers = get_headers_from_row(ws_target, header_row)
    if not target_headers:
        raise ValueError(f"No headers found in target sheet '{target_sheet}' at row {header_row}")

    # Normalize source column names
    source_columns = [str(col).strip() for col in df_src.columns]
    df_src.columns = source_columns

    # Find common columns in source order
    common_columns = [col for col in source_columns if col in target_headers]
    missing_in_target = [col for col in source_columns if col not in target_headers]

    if not common_columns:
        raise ValueError("No matching columns found between source sheet and target sheet.")

    print(f"Matching columns: {len(common_columns)}")
    if missing_in_target:
        print("Skipping columns not found in target:")
        for col in missing_in_target:
            print(f"  - {col}")

    # Keep only common columns, in source order
    df_to_copy = df_src[common_columns]

    # Map target columns once
    target_col_indexes = [target_headers[col] for col in common_columns]

    # Existing used rows in target before overwrite
    old_max_row = ws_target.max_row
    new_row_count = len(df_to_copy)
    new_last_row = data_start_row + new_row_count - 1 if new_row_count > 0 else data_start_row - 1

    print("\nWriting data to target sheet...")
    rows_written = 0

    # Fast row iteration from DataFrame
    data_rows = list(df_to_copy.itertuples(index=False, name=None))

    for row_offset, row_values in enumerate(
        tqdm(data_rows, total=len(data_rows), desc="Copying rows")
    ):
        excel_row = data_start_row + row_offset
        for col_idx, value in zip(target_col_indexes, row_values):
            ws_target.cell(row=excel_row, column=col_idx).value = value
        rows_written += 1

    # Clear leftover old rows if new dataset is smaller than old target data
    if old_max_row > new_last_row:
        clear_start = max(data_start_row, new_last_row + 1)
        clear_end = old_max_row
        print("\nClearing leftover old rows...")
        clear_extra_rows(
            ws_target,
            start_row=clear_start,
            end_row=clear_end,
            col_indexes=target_col_indexes,
            desc="Clearing old rows",
        )

    print("\nSaving target workbook...")
    wb_target.save(target_file)

    print("\nDone.")
    print(f"Rows copied      : {rows_written}")
    print(f"Columns copied   : {len(common_columns)}")
    print(f"Updated file     : {target_file}")


if __name__ == "__main__":
    SOURCE_FILE = r"C:\path\to\first.xlsx"
    TARGET_FILE = r"C:\path\to\second.xlsx"

    copy_sheet_data_pandas_openpyxl(
        source_file=SOURCE_FILE,
        target_file=TARGET_FILE,
        source_sheet="SheetA",
        target_sheet="SheetB",
        header_row=2,
        data_start_row=3,
    )