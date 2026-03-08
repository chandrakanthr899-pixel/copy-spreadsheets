import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def log(message: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def backup_file(path: str) -> str:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    base = os.path.basename(path)
    name, ext = os.path.splitext(base)

    ts = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{name}-{ts}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_headers_with_openpyxl(file_path: str, sheet_name: str, header_row: int) -> dict:
    """
    Returns:
        {
            "Header Name": 1-based Excel column index
        }
    """
    log(f"Reading headers from '{sheet_name}' in {os.path.basename(file_path)}")

    wb = load_workbook(file_path, read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        ws = wb[sheet_name]
        headers = {}

        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            key = normalize_header(val)
            if key and key not in headers:
                headers[key] = col

        log(f"Found {len(headers)} headers in '{sheet_name}'")
        return headers
    finally:
        wb.close()


def load_source_dataframe(src_path: str, src_sheet: str, header_row: int) -> pd.DataFrame:
    log(f"Loading source sheet '{src_sheet}' with pandas")
    df = pd.read_excel(
        src_path,
        sheet_name=src_sheet,
        header=header_row - 1,   # pandas is 0-based
        dtype=object,
        engine="openpyxl",
    )
    df.columns = [normalize_header(col) for col in df.columns]
    log(f"Loaded source rows: {len(df)} | columns: {len(df.columns)}")
    return df


def copy_sheet_data_pandas(
    src_path: str,
    dst_path: str,
    src_sheet: str = "SheetA",
    dst_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
):
    log("Starting Excel copy operation")

    # 1) Backup first
    log("Creating backups")
    b1 = backup_file(src_path)
    b2 = backup_file(dst_path)
    log(f"Backup created: {b1}")
    log(f"Backup created: {b2}")

    # 2) Read headers using openpyxl so we know exact Excel column positions
    src_headers = read_headers_with_openpyxl(src_path, src_sheet, header_row)
    dst_headers = read_headers_with_openpyxl(dst_path, dst_sheet, header_row)

    if not src_headers:
        raise ValueError(f"No headers found in source sheet '{src_sheet}' at row {header_row}")
    if not dst_headers:
        raise ValueError(f"No headers found in destination sheet '{dst_sheet}' at row {header_row}")

    # 3) Find only common columns
    common_cols = [col for col in src_headers if col in dst_headers]
    src_only_cols = [col for col in src_headers if col not in dst_headers]
    dst_only_cols = [col for col in dst_headers if col not in src_headers]

    log(f"Source columns found      : {len(src_headers)}")
    log(f"Destination columns found : {len(dst_headers)}")
    log(f"Common columns to update  : {len(common_cols)}")

    if not common_cols:
        raise ValueError("No matching column names found between source and destination.")

    if src_only_cols:
        log("Source-only columns will be skipped:")
        for col in src_only_cols[:20]:
            log(f"  - {col}")
        if len(src_only_cols) > 20:
            log(f"  ... and {len(src_only_cols) - 20} more")

    if dst_only_cols:
        log("Destination-only columns will be left untouched:")
        for col in dst_only_cols[:20]:
            log(f"  - {col}")
        if len(dst_only_cols) > 20:
            log(f"  ... and {len(dst_only_cols) - 20} more")

    # 4) Load source with pandas
    df_source = load_source_dataframe(src_path, src_sheet, header_row)

    # Keep only common columns, in source order
    df_common = df_source[[col for col in df_source.columns if col in common_cols]].copy()

    # Remove fully empty rows from what will actually be written
    before = len(df_common)
    df_common = df_common.dropna(how="all").reset_index(drop=True)
    removed = before - len(df_common)

    if removed:
        log(f"Removed {removed} fully empty source row(s)")

    if df_common.empty:
        log("No source data rows found to copy after filtering")
        return

    log(f"Rows to write: {len(df_common)}")

    # 5) Optional: clear old values only in matching destination columns
    # This avoids stale data below the new row count in common columns.
    log("Loading destination workbook for cleanup")
    wb = load_workbook(dst_path)
    try:
        if dst_sheet not in wb.sheetnames:
            raise ValueError(f"Destination sheet '{dst_sheet}' not found in {dst_path}")

        ws = wb[dst_sheet]
        max_row = ws.max_row

        log(
            f"Clearing old values only in matching destination columns "
            f"from row {data_start_row} to row {max_row}"
        )

        for idx, col_name in enumerate(common_cols, start=1):
            excel_col = dst_headers[col_name]
            col_letter = get_column_letter(excel_col)

            for row_num in range(data_start_row, max_row + 1):
                ws.cell(row=row_num, column=excel_col).value = None

            log(f"Cleared column {idx}/{len(common_cols)}: {col_name} ({col_letter})")

        wb.save(dst_path)
        log("Cleanup save completed")
    finally:
        wb.close()

    # 6) Write each common column back using pandas overlay
    # Writing one column at a time ensures untouched columns remain untouched.
    log("Writing updated data into destination sheet using pandas overlay")

    with pd.ExcelWriter(
        dst_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
    ) as writer:
        for idx, col_name in enumerate(common_cols, start=1):
            dst_excel_col_index = dst_headers[col_name]  # 1-based Excel index
            startcol_zero_based = dst_excel_col_index - 1

            log(
                f"Writing column {idx}/{len(common_cols)}: "
                f"'{col_name}' -> destination column {get_column_letter(dst_excel_col_index)}"
            )

            # Write just this one column, below the header rows
            df_common[[col_name]].to_excel(
                writer,
                sheet_name=dst_sheet,
                index=False,
                header=False,
                startrow=data_start_row - 1,   # pandas is 0-based
                startcol=startcol_zero_based,
            )

    log("Write completed successfully")
    log(
        f"Done. Data copied from '{src_sheet}' ({src_path}) "
        f"to '{dst_sheet}' ({dst_path})"
    )


if __name__ == "__main__":
    SRC_FILE = r"source.xlsx"
    DST_FILE = r"destination.xlsx"

    copy_sheet_data_pandas(
        src_path=SRC_FILE,
        dst_path=DST_FILE,
        src_sheet="July Template",
        dst_sheet="template",
        header_row=2,
        data_start_row=3,
    )