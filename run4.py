import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm


def backup_file(file_path: str) -> str:
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    folder = os.path.dirname(os.path.abspath(file_path))
    base_name = os.path.basename(file_path)
    name, ext = os.path.splitext(base_name)

    timestamp = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{name}-{timestamp}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(file_path, backup_path)
    return backup_path


def get_sheet_headers(ws, header_row: int) -> dict:
    """
    Returns a dict of:
        {header_name: column_index}
    using the given header row.
    """
    headers = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is None:
            continue
        header = str(cell.value).strip()
        if header:
            headers[header] = idx
    return headers


def clear_destination_data(ws, start_row: int, max_col: int):
    """
    Clears only cell values from start_row onward.
    Keeps formatting, formulas above the data region, column widths, styles, etc.
    """
    max_row = ws.max_row
    if max_row < start_row:
        return

    total_rows = max_row - start_row + 1
    for row in tqdm(
        range(start_row, max_row + 1),
        total=total_rows,
        desc="Clearing destination rows",
        unit="row"
    ):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None


def main(
    source_file: str,
    target_file: str,
    source_sheet: str = "SheetA",
    target_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
):
    print("Starting backup...")
    for file_path in tqdm([source_file, target_file], desc="Backing up files", unit="file"):
        backup_path = backup_file(file_path)
        print(f"Backup created: {backup_path}")

    print("\nLoading source data with pandas...")
    with tqdm(total=1, desc="Reading source sheet", unit="step") as pbar:
        # header_row=2 in Excel => header=1 in pandas (0-based)
        df_source = pd.read_excel(
            source_file,
            sheet_name=source_sheet,
            header=header_row - 1,
            engine="openpyxl"
        )
        pbar.update(1)

    # Drop fully empty rows
    df_source = df_source.dropna(how="all")

    print("\nOpening target workbook with openpyxl...")
    with tqdm(total=1, desc="Loading target workbook", unit="step") as pbar:
        wb_target = load_workbook(target_file)
        pbar.update(1)

    if target_sheet not in wb_target.sheetnames:
        raise ValueError(f"Target sheet '{target_sheet}' not found in {target_file}")

    ws_target = wb_target[target_sheet]

    print("\nReading target headers...")
    with tqdm(total=1, desc="Reading target header row", unit="step") as pbar:
        target_headers = get_sheet_headers(ws_target, header_row)
        pbar.update(1)

    if not target_headers:
        raise ValueError(f"No headers found in target sheet '{target_sheet}' at row {header_row}")

    # Normalize source columns
    source_columns = [str(col).strip() for col in df_source.columns]
    df_source.columns = source_columns

    common_columns = [col for col in source_columns if col in target_headers]

    if not common_columns:
        raise ValueError("No matching columns found between source and target sheets.")

    skipped_columns = [col for col in source_columns if col not in target_headers]
    if skipped_columns:
        print("\nThese source columns do not exist in the target sheet and will be skipped:")
        for col in skipped_columns:
            print(f" - {col}")

    # Keep only matching columns
    df_to_write = df_source[common_columns].copy()

    print(f"\nRows to copy: {len(df_to_write)}")
    print(f"Matching columns: {len(common_columns)}")

    # Clear existing target data only across used target columns
    max_target_col = ws_target.max_column
    clear_destination_data(ws_target, data_start_row, max_target_col)

    print("\nWriting data to target sheet...")
    total_rows = len(df_to_write)

    for row_offset, row_values in enumerate(
        tqdm(df_to_write.itertuples(index=False, name=None), total=total_rows, desc="Writing rows", unit="row"),
        start=0
    ):
        excel_row = data_start_row + row_offset
        row_dict = dict(zip(common_columns, row_values))

        for col_name, value in row_dict.items():
            target_col_idx = target_headers[col_name]
            ws_target.cell(row=excel_row, column=target_col_idx, value=value)

    print("\nSaving updated target workbook...")
    with tqdm(total=1, desc="Saving workbook", unit="step") as pbar:
        wb_target.save(target_file)
        pbar.update(1)

    print("\nDone.")
    print(f"Copied data from '{source_sheet}' in '{source_file}'")
    print(f"to '{target_sheet}' in '{target_file}'")


if __name__ == "__main__":
    SOURCE_FILE = r"C:\path\to\first.xlsx"
    TARGET_FILE = r"C:\path\to\second.xlsx"

    main(
        source_file=SOURCE_FILE,
        target_file=TARGET_FILE,
        source_sheet="SheetA",
        target_sheet="SheetB",
        header_row=2,
        data_start_row=3,
    )