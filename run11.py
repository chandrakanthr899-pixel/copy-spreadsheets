import os
import shutil
import logging
from datetime import datetime

from openpyxl import load_workbook


# ----------------------------
# Logging setup
# ----------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)


def log(message: str):
    logging.info(message)
    print(message, flush=True)


def backup_file(path: str) -> str:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")

    folder = os.path.dirname(os.path.abspath(path))
    base = os.path.basename(path)
    name, ext = os.path.splitext(base)

    ts = datetime.now().strftime("%H%M%S")
    backup_name = f"backup-{name}-{ts}{ext}"
    backup_path = os.path.join(folder, backup_name)

    shutil.copy2(path, backup_path)
    return backup_path


def read_headers(ws, header_row: int) -> dict:
    """
    Returns dict: {header_name: column_index}
    Ignores empty header cells.
    """
    log(f"Reading headers from row {header_row} in sheet '{ws.title}'")

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            headers[key] = col

    log(f"Found {len(headers)} headers in sheet '{ws.title}'")
    return headers


def clear_data_rows_fast(ws, start_row: int):
    """
    Faster clearing by deleting all rows from start_row onward.
    This is much faster than clearing cell-by-cell.
    """
    max_row = ws.max_row
    if max_row < start_row:
        log(f"No rows to clear in destination sheet '{ws.title}'")
        return

    rows_to_delete = max_row - start_row + 1
    log(
        f"Deleting {rows_to_delete} existing data row(s) "
        f"from sheet '{ws.title}' starting at row {start_row}"
    )
    ws.delete_rows(start_row, rows_to_delete)
    log("Finished deleting old destination rows")


def build_source_row_generator(
    src_ws,
    src_header_indexes: list,
    common_cols: list,
    data_start_row: int,
    last_row: int,
    batch_size: int = 1000
):
    """
    Streams source rows as dicts for only the common columns.
    Adds status logging every batch while reading.
    """
    batch_count = 0

    for src_r in range(data_start_row, last_row + 1):
        row_data = {}

        for h, src_c in zip(common_cols, src_header_indexes):
            row_data[h] = src_ws.cell(row=src_r, column=src_c).value

        yield row_data

        if (src_r - data_start_row + 1) % batch_size == 0:
            batch_count += 1
            log(
                f"Read batch {batch_count}: source rows "
                f"{src_r - batch_size + 1} to {src_r}"
            )


def write_batches_to_destination(
    dst_ws,
    dst_header_indexes: list,
    source_rows,
    data_start_row: int,
    total_rows: int,
    batch_size: int = 1000
):
    """
    Writes rows to destination in batches and logs after every batch.
    """
    log(
        f"Starting batch write to destination sheet '{dst_ws.title}' "
        f"for {total_rows} row(s) with batch size {batch_size}"
    )

    batch = []
    current_dst_row = data_start_row
    batch_num = 0
    processed = 0

    for row_data in source_rows:
        batch.append(row_data)

        if len(batch) >= batch_size:
            batch_num += 1
            batch_start_row = current_dst_row

            for row_dict in batch:
                for dst_c, value in zip(dst_header_indexes, row_dict.values()):
                    dst_ws.cell(row=current_dst_row, column=dst_c).value = value
                current_dst_row += 1

            processed += len(batch)
            batch_end_row = current_dst_row - 1
            log(
                f"Wrote batch {batch_num}: destination rows {batch_start_row} to {batch_end_row} "
                f"({processed}/{total_rows} rows completed)"
            )
            batch = []

    if batch:
        batch_num += 1
        batch_start_row = current_dst_row

        for row_dict in batch:
            for dst_c, value in zip(dst_header_indexes, row_dict.values()):
                dst_ws.cell(row=current_dst_row, column=dst_c).value = value
            current_dst_row += 1

        processed += len(batch)
        batch_end_row = current_dst_row - 1
        log(
            f"Wrote final batch {batch_num}: destination rows {batch_start_row} to {batch_end_row} "
            f"({processed}/{total_rows} rows completed)"
        )

    log("Finished writing all batches to destination sheet")


def copy_sheet_data(
    src_path: str,
    dst_path: str,
    src_sheet: str = "SheetA",
    dst_sheet: str = "SheetB",
    header_row: int = 2,
    data_start_row: int = 3,
    batch_size: int = 1000,
):
    log("===== STARTING EXCEL COPY PROCESS =====")

    # 1) Backups FIRST
    log("Creating backups before making any changes")
    b1 = backup_file(src_path)
    b2 = backup_file(dst_path)
    log(f"Backup created: {b1}")
    log(f"Backup created: {b2}")

    # 2) Load workbooks
    log(f"Loading source workbook in read-only mode: {src_path}")
    src_wb = load_workbook(src_path, data_only=False, read_only=True)

    log(f"Loading destination workbook in normal mode: {dst_path}")
    dst_wb = load_workbook(dst_path, data_only=False)

    try:
        if src_sheet not in src_wb.sheetnames:
            raise ValueError(f"Source sheet '{src_sheet}' not found in {src_path}")
        if dst_sheet not in dst_wb.sheetnames:
            raise ValueError(f"Destination sheet '{dst_sheet}' not found in {dst_path}")

        log(f"Accessing source sheet: '{src_sheet}'")
        src_ws = src_wb[src_sheet]

        log(f"Accessing destination sheet: '{dst_sheet}'")
        dst_ws = dst_wb[dst_sheet]

        # 3) Header maps
        src_headers = read_headers(src_ws, header_row)
        dst_headers = read_headers(dst_ws, header_row)

        if not src_headers:
            raise ValueError(f"No headers found in source sheet '{src_sheet}' at row {header_row}")
        if not dst_headers:
            raise ValueError(f"No headers found in destination sheet '{dst_sheet}' at row {header_row}")

        # 4) Common columns only
        common_cols = [h for h in src_headers.keys() if h in dst_headers]
        missing_in_dst = [h for h in src_headers.keys() if h not in dst_headers]

        log(f"Source header count      : {len(src_headers)}")
        log(f"Destination header count : {len(dst_headers)}")
        log(f"Common header count      : {len(common_cols)}")

        if not common_cols:
            raise ValueError("No matching column names found between source and destination.")

        if missing_in_dst:
            log("These source columns are missing in destination and will be skipped:")
            for h in missing_in_dst:
                log(f"  - {h}")

        # Precompute column indexes once for speed
        src_header_indexes = [src_headers[h] for h in common_cols]
        dst_header_indexes = [dst_headers[h] for h in common_cols]

        # 5) Determine source data range
        last_row = src_ws.max_row
        if last_row < data_start_row:
            log("No data rows found to copy (source has no rows below header).")
            return

        total_rows = last_row - data_start_row + 1
        log(f"Source data rows detected: {total_rows} (from row {data_start_row} to {last_row})")

        # 6) Fast clear destination rows
        clear_data_rows_fast(dst_ws, data_start_row)

        # 7) Build source row generator
        log("Preparing source row generator for batch processing")
        source_rows = build_source_row_generator(
            src_ws=src_ws,
            src_header_indexes=src_header_indexes,
            common_cols=common_cols,
            data_start_row=data_start_row,
            last_row=last_row,
            batch_size=batch_size
        )

        # 8) Write in batches
        write_batches_to_destination(
            dst_ws=dst_ws,
            dst_header_indexes=dst_header_indexes,
            source_rows=source_rows,
            data_start_row=data_start_row,
            total_rows=total_rows,
            batch_size=batch_size
        )

        # 9) Save destination workbook
        log(f"Saving destination workbook: {dst_path}")
        dst_wb.save(dst_path)
        log("Destination workbook saved successfully")

    finally:
        log("Closing workbooks")
        try:
            src_wb.close()
        except Exception:
            pass
        try:
            dst_wb.close()
        except Exception:
            pass

    log("===== COPY PROCESS COMPLETED SUCCESSFULLY =====")


if __name__ == "__main__":
    SRC_FILE = r"source.xlsx"
    DST_FILE = r"destination.xlsx"

    copy_sheet_data(
        src_path=SRC_FILE,
        dst_path=DST_FILE,
        src_sheet="July Template",
        dst_sheet="template",
        header_row=2,
        data_start_row=3,
        batch_size=1000,   # try 500 / 1000 / 2000
    )